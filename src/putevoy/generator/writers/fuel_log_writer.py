"""Дописывает в Топливо.xlsx новый месяц по инструкции §6."""
from __future__ import annotations

from copy import copy
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.cell.cell import Cell
from openpyxl.utils import get_column_letter

from ..calendar import MONTH_NOMINATIVE_RU
from ..models import MonthlyOutput

COL_DATE = 4         # D
COL_LITERS = 5       # E
COL_PRICE = 6        # F
COL_SUM = 7          # G
COL_FROM = 8         # H
COL_TO = 9           # I
COL_ODO_BEFORE = 10  # J
COL_ODO_AFTER = 11   # K
COL_KM = 12          # L
COL_FUEL_BAL = 13    # M
COL_CONSUMPTION = 14 # N


def _find_last_data_row(ws) -> int:
    """Найти последнюю строку, где есть хоть что-то значимое (D..N)."""
    last = 1
    for r in range(ws.max_row, 0, -1):
        for c in range(COL_DATE, COL_CONSUMPTION + 1):
            if ws.cell(r, c).value not in (None, ""):
                return r
        last = r
    return last


def _clone_style_from(template_cell: Cell, target_cell: Cell) -> None:
    if template_cell.has_style:
        target_cell.font = copy(template_cell.font)
        target_cell.fill = copy(template_cell.fill)
        target_cell.border = copy(template_cell.border)
        target_cell.alignment = copy(template_cell.alignment)
        target_cell.number_format = template_cell.number_format


def append_month(
    template_path: Path,
    output_path: Path,
    out: MonthlyOutput,
) -> Path:
    wb = load_workbook(template_path)
    ws = wb.active

    last_row = _find_last_data_row(ws)
    # 2 пустые строки → строка заголовка → пустая → данные
    header_row = last_row + 3
    first_data_row = header_row + 2

    month_label = f"{MONTH_NOMINATIVE_RU[out.input.month]} {out.input.year}"
    ws.cell(header_row, COL_DATE).value = month_label

    cur_row = first_data_row
    first_data_row_recorded = cur_row

    for day in out.days:
        # Группируем строки по поездкам: каждая поездка = (outbound, inbound)
        # У нас day.trips уже идут парами outbound, inbound по trip_index
        for trip_idx in range(day.trip_count):
            outbound = day.trips[trip_idx * 2]
            inbound = day.trips[trip_idx * 2 + 1]

            # OUTBOUND
            r = cur_row
            # Дата — только в самой первой строке дня
            if trip_idx == 0:
                ws.cell(r, COL_DATE).value = day.date
                ws.cell(r, COL_DATE).number_format = "yyyy-mm-dd"
            # Заправка — в первой строке дня, если есть
            if day.fueling and trip_idx == 0:
                ws.cell(r, COL_LITERS).value = day.fueling.liters
                ws.cell(r, COL_PRICE).value = day.fueling.price_per_l
                ws.cell(r, COL_SUM).value = f"=E{r}*F{r}"
            ws.cell(r, COL_FROM).value = outbound.from_address
            ws.cell(r, COL_TO).value = outbound.to_address
            if r == first_data_row_recorded:
                # Для самой первой строки месяца — абсолютное число (см. инструкцию)
                ws.cell(r, COL_ODO_BEFORE).value = day.odometer_start
            else:
                ws.cell(r, COL_ODO_BEFORE).value = f"=K{r - 1}"
            ws.cell(r, COL_ODO_AFTER).value = f"=J{r}+L{r}"
            ws.cell(r, COL_KM).value = outbound.km
            # Остаток: =E+M_пред−N (если есть заправка) или =M_пред−N
            if r == first_data_row_recorded:
                # Первая строка: явное число для остатка до = carry_over + заправка - расход
                # Чтобы не плодить разные формулы, используем число
                start = out.input.carry_over.last_fuel_balance_l
                if day.fueling:
                    ws.cell(r, COL_FUEL_BAL).value = (
                        start + day.fueling.liters - outbound.consumption_l
                    )
                else:
                    ws.cell(r, COL_FUEL_BAL).value = start - outbound.consumption_l
            else:
                if day.fueling and trip_idx == 0:
                    ws.cell(r, COL_FUEL_BAL).value = f"=(E{r}+M{r - 1})-N{r}"
                else:
                    ws.cell(r, COL_FUEL_BAL).value = f"=M{r - 1}-N{r}"
            ws.cell(r, COL_CONSUMPTION).value = outbound.consumption_l

            cur_row += 1

            # INBOUND
            r = cur_row
            ws.cell(r, COL_FROM).value = inbound.from_address
            ws.cell(r, COL_TO).value = inbound.to_address
            ws.cell(r, COL_ODO_BEFORE).value = f"=K{r - 1}"
            ws.cell(r, COL_ODO_AFTER).value = f"=J{r}+L{r}"
            ws.cell(r, COL_KM).value = inbound.km
            ws.cell(r, COL_FUEL_BAL).value = f"=M{r - 1}-N{r}"
            ws.cell(r, COL_CONSUMPTION).value = inbound.consumption_l

            cur_row += 1

    # Итого:
    total_row = cur_row
    ws.cell(total_row, COL_DATE).value = "Итого:"
    ws.cell(total_row, COL_KM).value = f"=SUM(L{first_data_row}:L{cur_row - 1})"

    # Скопируем стиль из любой подходящей соседней ячейки заголовка/данных,
    # если она есть в шаблоне (стилистика последних месяцев похожа)
    if last_row > 1:
        for c in range(COL_DATE, COL_CONSUMPTION + 1):
            ref_cell = ws.cell(last_row, c)
            if ref_cell.has_style:
                for rr in range(header_row, total_row + 1):
                    _clone_style_from(ref_cell, ws.cell(rr, c))

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    return output_path
