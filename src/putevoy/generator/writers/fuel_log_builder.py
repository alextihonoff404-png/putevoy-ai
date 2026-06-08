"""Сборка Топливо.xlsx из истории прогонов в БД (без шаблонного файла).

Структура (как в инструкции §6):
  Заголовок столбцов (D..N): Дата | Кол-во, л | Стоимость, руб. | Сумма |
      Отправление | Назначение | Одометр до | Одометр после | Пройдено |
      Остаток бензина | Расход топлива
  Для каждого месяца:
    2 пустые строки
    Заголовок «Январь 2026» в колонке D
    Пустая строка
    Для каждого дня — по 2 строки (туда/обратно) на каждую поездку
    Итого: с SUM(L)
"""
from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from ..calendar import MONTH_NOMINATIVE_RU
from ..models import MonthlyOutput

HEADER_COLS = [
    "Дата", "Кол-во, л", "Стоимость, руб.", "Сумма",
    "Отправление", "Назначение", "Одометр до, км", "Одометр после, км",
    "Пройдено, км", "Остаток бензина", "Расход топлива, л",
]
START_COL = 4  # D
COL_COUNT = len(HEADER_COLS)


def _write_header(ws, row: int = 1) -> None:
    bold = Font(bold=True)
    fill = PatternFill("solid", fgColor="E0E7EF")
    for i, title in enumerate(HEADER_COLS):
        c = ws.cell(row, START_COL + i, value=title)
        c.font = bold
        c.fill = fill
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _write_month(
    ws, start_row: int, out: MonthlyOutput,
    carry_over_balance_l: float,
    prev_month_last_row: int | None = None,
) -> tuple[int, int]:
    """Записать один месяц начиная с start_row.

    Args:
        prev_month_last_row: номер последней data-строки предыдущего месяца
            в этом же файле. Если задан — первая строка месяца будет ссылаться
            формулой на M{prev} и K{prev}, что делает преемственность явной
            (и Excel пересчитывает значения, если предыдущий месяц поправили).
            Если None — берётся абсолютное число `carry_over_balance_l`
            (вариант для самого первого месяца в файле).

    Returns:
        (next_free_row, last_data_row_of_this_month)
    """
    month_label = f"{MONTH_NOMINATIVE_RU[out.input.month]} {out.input.year}"
    header_row = start_row + 2  # 2 пустые
    ws.cell(header_row, START_COL, value=month_label).font = Font(bold=True)
    data_start = header_row + 2  # пустая + начало данных

    r = data_start
    last_data_row_of_month = data_start
    is_first_data_row_of_month = True
    for day in out.days:
        for trip_idx in range(day.trip_count):
            outbound = day.trips[trip_idx * 2]
            inbound = day.trips[trip_idx * 2 + 1]

            # OUTBOUND
            if trip_idx == 0:
                ws.cell(r, 4, value=day.date).number_format = "yyyy-mm-dd"
            if day.fueling and trip_idx == 0:
                ws.cell(r, 5, value=day.fueling.liters)
                ws.cell(r, 6, value=day.fueling.price_per_l)
                ws.cell(r, 7, value=f"=E{r}*F{r}")
            ws.cell(r, 8, value=outbound.from_address)
            ws.cell(r, 9, value=outbound.to_address)

            # Одометр и остаток на первой строке месяца:
            # если предыдущий месяц есть — ссылаемся на него формулой,
            # иначе записываем абсолютное число (только в самом первом месяце файла).
            if is_first_data_row_of_month:
                if prev_month_last_row is not None:
                    ws.cell(r, 10, value=f"=K{prev_month_last_row}")
                else:
                    ws.cell(r, 10, value=day.odometer_start)
            else:
                ws.cell(r, 10, value=f"=K{r - 1}")
            ws.cell(r, 11, value=f"=J{r}+L{r}")
            ws.cell(r, 12, value=outbound.km)

            if is_first_data_row_of_month:
                if prev_month_last_row is not None:
                    # Формула со ссылкой на остаток конца прошлого месяца
                    if day.fueling:
                        ws.cell(r, 13, value=f"=(E{r}+M{prev_month_last_row})-N{r}")
                    else:
                        ws.cell(r, 13, value=f"=M{prev_month_last_row}-N{r}")
                else:
                    # Самый первый месяц — число
                    start = carry_over_balance_l
                    if day.fueling:
                        ws.cell(r, 13, value=round(start + day.fueling.liters - outbound.consumption_l, 4))
                    else:
                        ws.cell(r, 13, value=round(start - outbound.consumption_l, 4))
            else:
                if day.fueling and trip_idx == 0:
                    ws.cell(r, 13, value=f"=(E{r}+M{r - 1})-N{r}")
                else:
                    ws.cell(r, 13, value=f"=M{r - 1}-N{r}")
            ws.cell(r, 14, value=outbound.consumption_l)
            last_data_row_of_month = r
            r += 1
            is_first_data_row_of_month = False

            # INBOUND
            ws.cell(r, 8, value=inbound.from_address)
            ws.cell(r, 9, value=inbound.to_address)
            ws.cell(r, 10, value=f"=K{r - 1}")
            ws.cell(r, 11, value=f"=J{r}+L{r}")
            ws.cell(r, 12, value=inbound.km)
            ws.cell(r, 13, value=f"=M{r - 1}-N{r}")
            ws.cell(r, 14, value=inbound.consumption_l)
            last_data_row_of_month = r
            r += 1

    # Итого
    ws.cell(r, 4, value="Итого:").font = Font(bold=True)
    ws.cell(r, 12, value=f"=SUM(L{data_start}:L{r - 1})").font = Font(bold=True)
    return r + 1, last_data_row_of_month


def build_fuel_log(out_path: Path, months: list[MonthlyOutput],
                   initial_balance_l: float) -> Path:
    """Собрать полный накопительный Топливо.xlsx из списка прогонов.

    `initial_balance_l` — остаток на момент ДО первого месяца в списке.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Топливные отчеты"
    _write_header(ws, row=1)

    # Ширины колонок (приблизительно как в реальном файле)
    widths = [13, 11, 14, 11, 38, 38, 13, 13, 12, 14, 14]
    for i, w in enumerate(widths):
        ws.column_dimensions[get_column_letter(START_COL + i)].width = w

    current_row = 2
    running_balance = initial_balance_l
    prev_month_last_row: int | None = None
    for month_out in months:
        next_row, last_data_row = _write_month(
            ws, current_row, month_out, running_balance,
            prev_month_last_row=prev_month_last_row,
        )
        running_balance = month_out.final_state.last_fuel_balance_l
        prev_month_last_row = last_data_row
        current_row = next_row

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    return out_path
