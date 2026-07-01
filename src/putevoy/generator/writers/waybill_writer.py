"""Генерация .xlsx путевых листов на месяц.

Подход: открываем шаблон (xlsx с одной готовой парой лицевая+оборотная),
клонируем пару нужное количество раз, переименовываем и заполняем ячейки
по cell_map. Шаблон может быть пользовательский (премиум) или из пресета.
"""
from __future__ import annotations

import json
from copy import copy
from datetime import datetime, time
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.utils import column_index_from_string, get_column_letter
from openpyxl.worksheet.page import PageMargins

from ..calendar import MONTH_GENITIVE_RU
from ..models import GeneratedDay, MonthlyOutput

CELL_MAP_DIR = Path(__file__).resolve().parents[2] / "cell_maps"


def _load_cell_map(name: str) -> dict:
    return json.loads((CELL_MAP_DIR / f"{name}.json").read_text(encoding="utf-8"))


def _set(ws, coord_or_row, value, col=None):
    """Безопасно записать значение в ячейку: если попали в MergedCell — найти top-left объединения.

    Поддерживает оба вызова: _set(ws, 'A1', val) и _set(ws, row, val, col=N).
    """
    if col is None:
        cell = ws[coord_or_row]
        if isinstance(cell, MergedCell):
            for mr in ws.merged_cells.ranges:
                if cell.coordinate in mr:
                    ws.cell(mr.min_row, mr.min_col).value = value
                    return
            return  # на всякий случай: не пишем
        cell.value = value
    else:
        cell = ws.cell(coord_or_row, col)
        if isinstance(cell, MergedCell):
            for mr in ws.merged_cells.ranges:
                if cell.coordinate in mr:
                    ws.cell(mr.min_row, mr.min_col).value = value
                    return
            return
        cell.value = value


def _pair_sheet_name(d1, d2, month_short: str) -> str:
    return f"{d1.day:02d},{d2.day:02d} {month_short}"


def _single_sheet_name(d, month_short: str) -> str:
    return f"{d.day:02d} {month_short}"


def _format_dt(dt: datetime) -> str:
    return dt.strftime("%d.%m.%Y %H:%M")


def _format_d(d) -> str:
    return d.strftime("%d.%m.%Y")


def _fill_front_day(ws, day_map: dict, day: GeneratedDay, waybill_number: int) -> None:
    if "waybill_number" in day_map:
        _set(ws, day_map["waybill_number"], waybill_number)
    _set(ws, day_map["day_of_month"], day.date.day)
    _set(ws, day_map["month_genitive"], MONTH_GENITIVE_RU[day.date.month])
    _set(ws, day_map["year_short"], day.date.strftime("%y"))
    _set(ws, day_map["medical_exam_date_top"], _format_d(day.date))
    _set(ws, day_map["medical_exam_date_bottom"], _format_d(day.date))
    _set(ws, day_map["release_datetime"], _format_dt(day.release_datetime))
    _set(ws, day_map["return_datetime"], _format_dt(day.return_datetime))
    _set(ws, day_map["odometer_end"], day.odometer_end)
    if day.fueling:
        _set(ws, day_map["fuel_issued"], round(day.fueling.liters, 2))
    else:
        _set(ws, day_map["fuel_issued"], None)
    _set(ws, day_map["fuel_balance_end"], round(day.fuel_balance_end, 2))
    _set(ws, day_map["fuel_consumption_formula"], day_map["fuel_consumption_formula_text"])


def _fill_front_day_left(ws, day_map: dict, day: GeneratedDay, waybill_number: int) -> None:
    _fill_front_day(ws, day_map, day, waybill_number)
    _set(ws, day_map["odometer_start"], day.odometer_start)
    _set(ws, day_map["fuel_balance_start"], round(day.fuel_balance_start, 2))


def _fill_front_day_right(ws, day_map: dict, day: GeneratedDay, waybill_number: int) -> None:
    _fill_front_day(ws, day_map, day, waybill_number)
    _set(ws, day_map["odometer_start"], day_map["odometer_start_formula_text"])
    _set(ws, day_map["fuel_balance_start"], day_map["fuel_balance_start_formula_text"])


def _fill_back_half(ws, half_map: dict, day: GeneratedDay) -> None:
    from_col = column_index_from_string(half_map["from_col"])
    to_col = column_index_from_string(half_map["to_col"])
    km_col = column_index_from_string(half_map["km_col"])
    dep_col = column_index_from_string(half_map["depart_col"])
    arr_col = column_index_from_string(half_map["arrive_col"])
    start_row = half_map["start_row"]
    row_step = half_map["row_step"]

    for i, trip in enumerate(day.trips):
        r = start_row + i
        _set(ws, r, trip.from_address, col=from_col)
        _set(ws, r, trip.to_address, col=to_col)
        _set(ws, r, trip.km, col=km_col)
        _set(ws, r, trip.depart, col=dep_col)
        _set(ws, r, trip.arrive, col=arr_col)

    total_km = sum(t.km for t in day.trips)
    _set(ws, half_map["total_km_cell"], total_km)


def _clear_back_half(ws, half_map: dict, rows_to_clear: int = 14) -> None:
    cols = [
        column_index_from_string(half_map[k])
        for k in ("from_col", "to_col", "km_col", "depart_col", "arrive_col")
    ]
    start_row = half_map["start_row"]
    for offset in range(rows_to_clear):
        for c in cols:
            _set(ws, start_row + offset, None, col=c)
    _set(ws, half_map["total_km_cell"], None)


def _organization_full_line(org) -> str:
    """Строка вида «Название, ОГРН ..., адрес, телефон» для шапки формы 390н."""
    parts = [org.name]
    if org.ogrn:
        parts.append(f"ОГРН {org.ogrn}")
    if org.address:
        parts.append(org.address)
    if org.phone:
        parts.append(org.phone)
    return ", ".join(parts)


def _fill_header(ws, header_map: dict, out: MonthlyOutput) -> None:
    """Заполнить статическую шапку документа (организация/водитель/ТС) —
    одна и та же для обеих половин листа и для всех дней месяца.
    """
    org = out.input.organization
    driver = out.input.driver
    vehicle = out.input.vehicle
    license_text = (
        f"{driver.license_number} от {driver.license_issue_date:%d.%m.%Y}г."
    )
    values = {
        "organization_full_line": _organization_full_line(org),
        # В блок «Задание водителю» — сокращённое имя (место ограничено),
        # с фолбэком на полное для профилей без короткого наименования.
        "organization_short_name": org.short_name or org.name,
        "mechanic_name": org.mechanic_name,
        "driver_full_name": driver.full_name,
        "driver_license": license_text,
        "driver_snils": driver.snils,
        "vehicle_make_model": vehicle.make_model,
        "vehicle_license_plate": vehicle.license_plate,
        "vehicle_fuel_grade": vehicle.fuel_grade,
    }
    for key, value in values.items():
        for coord in header_map.get(key, []):
            _set(ws, coord, value)


def write_waybills(
    template_path: Path,
    output_path: Path,
    out: MonthlyOutput,
    cell_map_name: str = "preset_mintrans_390_v1",
) -> Path:
    cm = _load_cell_map(cell_map_name)
    month_short = cm["sheet_naming"]["month_short_genitive"][str(out.input.month)]

    wb = load_workbook(template_path)

    # Шаблонные листы (первая пара) — берём как образец
    template_front = wb.worksheets[0]
    template_back = wb.worksheets[1] if len(wb.worksheets) > 1 else None
    if template_back is None or not template_back.title.endswith(" 2"):
        raise ValueError("Шаблон должен содержать хотя бы одну пару: лицевая + оборотная (' 2')")

    template_front_name = template_front.title
    template_back_name = template_back.title

    # Разбиваем рабочие дни на пары
    days = out.days
    pairs: list[tuple[GeneratedDay, GeneratedDay | None]] = []
    i = 0
    while i < len(days):
        pairs.append((days[i], days[i + 1] if i + 1 < len(days) else None))
        i += 2

    # Создаём нужное число копий пары шаблона (уже одна есть)
    needed_pairs = len(pairs)
    existing_pairs_in_template = len(wb.worksheets) // 2

    new_pair_sheets: list[tuple[object, object]] = []

    # Используем существующие пары шаблона + добавляем недостающие копированием первой пары
    for p_idx in range(needed_pairs):
        front_idx = p_idx * 2
        back_idx = p_idx * 2 + 1
        if back_idx < len(wb.worksheets):
            new_pair_sheets.append((wb.worksheets[front_idx], wb.worksheets[back_idx]))
        else:
            new_front = wb.copy_worksheet(template_front)
            new_back = wb.copy_worksheet(template_back)
            new_pair_sheets.append((new_front, new_back))

    # Удаляем лишние пары шаблона (если шаблон содержал больше пар, чем нужно)
    used_titles = set()
    for front_ws, back_ws in new_pair_sheets:
        used_titles.add(front_ws.title)
        used_titles.add(back_ws.title)
    for sh in list(wb.worksheets):
        if sh.title not in used_titles:
            del wb[sh.title]

    # Заполняем — со сквозной нумерацией путевых листов через весь месяц
    for pair_idx, ((front_ws, back_ws), (d1, d2)) in enumerate(zip(new_pair_sheets, pairs)):
        num_left = pair_idx * 2 + 1
        num_right = pair_idx * 2 + 2

        # Имена листов
        if d2 is not None:
            pair_name = _pair_sheet_name(d1.date, d2.date, month_short)
            back_name = pair_name + cm["sheet_naming"]["back_suffix"]
        else:
            pair_name = _single_sheet_name(d1.date, month_short)
            back_name = pair_name + cm["sheet_naming"]["back_suffix"]

        # Переименовываем (с осторожностью к дублям)
        if front_ws.title != pair_name:
            front_ws.title = _unique_title(wb, pair_name, front_ws)
        if back_ws.title != back_name:
            back_ws.title = _unique_title(wb, back_name, back_ws)

        # Лицевая
        _fill_header(front_ws, cm["header"], out)
        _fill_front_day_left(front_ws, cm["front"]["day_left"], d1, num_left)
        if d2 is not None:
            _fill_front_day_right(front_ws, cm["front"]["day_right"], d2, num_right)
        else:
            _clear_front_day_right(front_ws, cm["front"]["day_right"])

        # Оборотная
        _clear_back_half(back_ws, cm["back"]["day_1_right_half"])
        _clear_back_half(back_ws, cm["back"]["day_2_left_half"])
        _fill_back_half(back_ws, cm["back"]["day_1_right_half"], d1)
        if d2 is not None:
            _fill_back_half(back_ws, cm["back"]["day_2_left_half"], d2)

    # Гарантируем одинаковые поля и центрирование на ВСЕХ листах книги,
    # чтобы двусторонняя печать давала совпадение лицевой и оборотной.
    for ws in wb.worksheets:
        ws.page_margins = PageMargins(
            left=0.5, right=0.5, top=0.5, bottom=0.5,
            header=0.3, footer=0.3,
        )
        ws.print_options.horizontalCentered = True
        ws.print_options.verticalCentered = False

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    return output_path


def _unique_title(wb, desired: str, sheet) -> str:
    """Безопасное переименование: если такое имя уже есть у другого листа — добавим suffix."""
    if desired not in wb.sheetnames or wb[desired] is sheet:
        return desired
    suffix = 1
    while f"{desired} ({suffix})" in wb.sheetnames:
        suffix += 1
    return f"{desired} ({suffix})"


def _clear_front_day_right(ws, day_map: dict) -> None:
    for k in [
        "waybill_number",
        "day_of_month", "month_genitive", "year_short",
        "medical_exam_date_top", "medical_exam_date_bottom",
        "release_datetime", "return_datetime",
        "odometer_start", "odometer_end",
        "fuel_issued", "fuel_balance_start", "fuel_balance_end",
        "fuel_consumption_formula",
    ]:
        if k in day_map:
            _set(ws, day_map[k], None)
