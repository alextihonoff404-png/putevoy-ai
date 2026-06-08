"""Извлечение стартового состояния (одометр, остаток, дата) из старого Топливо.xlsx.

Логика: ищем последнюю строку данных (не "Итого:" и не пустую) и читаем оттуда
K (одометр после), M (остаток топлива), последнее найденное D (дата) либо строку
выше — если в этой строке нет даты.
"""
from __future__ import annotations

from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path
from typing import Optional

from openpyxl import load_workbook


@dataclass
class ImportedState:
    odometer_km: int
    fuel_balance_l: float
    last_date: date


def import_state_from_xlsx(path: Path) -> Optional[ImportedState]:
    """Открыть файл, найти последние data-значения. Вернуть None если не удалось."""
    try:
        wb = load_workbook(path, data_only=True)
    except Exception:
        return None

    ws = wb.active
    last_data_row = None
    for r in range(ws.max_row, 1, -1):
        d_val = ws.cell(r, 4).value  # D
        if isinstance(d_val, str) and d_val.strip() == "Итого:":
            continue
        # Считаем строку «датной», если есть K (одометр после) или M (остаток)
        k_val = ws.cell(r, 11).value
        m_val = ws.cell(r, 13).value
        if k_val is not None or m_val is not None:
            last_data_row = r
            break

    if last_data_row is None:
        return None

    k_val = ws.cell(last_data_row, 11).value
    m_val = ws.cell(last_data_row, 13).value
    if k_val is None or m_val is None:
        return None

    # Дата: ищем D в этой или ближайшей выше строке
    last_date_val = None
    for r in range(last_data_row, max(0, last_data_row - 200), -1):
        d_val = ws.cell(r, 4).value
        if isinstance(d_val, datetime):
            last_date_val = d_val.date()
            break
        if isinstance(d_val, date) and not isinstance(d_val, datetime):
            last_date_val = d_val
            break

    if last_date_val is None:
        return None

    try:
        return ImportedState(
            odometer_km=int(k_val),
            fuel_balance_l=float(m_val),
            last_date=last_date_val,
        )
    except (TypeError, ValueError):
        return None
