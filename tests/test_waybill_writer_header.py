"""Регрессия: шапка путевого листа (организация/водитель/ТС) должна отражать
данные КОНКРЕТНОГО прогона, а не то, что было зашито в builtin-шаблоне.

История бага: builtin-шаблон был склонирован из реального путевого листа
разработчика и содержал его орг./ФИО/машину прямо в ячейках. write_waybills
заполнял только дни/пробег/топливо, а «шапку» никогда не трогал — поэтому
каждый пользователь видел чужие (мои) данные независимо от своего профиля.
"""
from __future__ import annotations

from datetime import date
from pathlib import Path

import openpyxl

from putevoy.generator.generate import generate_month_auto_seed
from putevoy.generator.models import (
    CarryOver, Driver, Fueling, MonthlyInput, Organization, Route, Vehicle,
)
from putevoy.generator.writers.waybill_writer import write_waybills

BUILTIN_TEMPLATE = (
    Path(__file__).resolve().parents[1]
    / "src" / "putevoy" / "builtin_templates" / "preset_mintrans_390_v1.xlsx"
)


def _input_for(org_name: str, driver_name: str, make_model: str, plate: str) -> MonthlyInput:
    return MonthlyInput(
        year=2026, month=8, seed=1,
        organization=Organization(
            name=org_name, mechanic_name="Механик Тестов",
            address="г. Тест, ул. Тестовая, д. 1", ogrn="1234567890123",
            phone="(000) 000-00-00",
        ),
        driver=Driver(
            full_name=driver_name, snils="111-222-333 44",
            license_number="77 00 №000000", license_issue_date=date(2022, 1, 1),
        ),
        vehicle=Vehicle(
            make_model=make_model, license_plate=plate,
            fuel_grade="АИ-92", tank_capacity_l=45.0,
            base_address="г. Тест, ул. Гаражная, д. 5",
        ),
        routes=[Route(address="Точка А", km_one_way=10, consumption_l_one_way=1.0)],
        fuelings=[Fueling(date=date(2026, 8, 3), liters=30, price_per_l=55, sum=1650)],
        carry_over=CarryOver(last_odometer_km=1000, last_fuel_balance_l=20, last_date=date(2026, 7, 31)),
    )


def _front_header_cells(xlsx_path: Path) -> dict:
    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb.worksheets[0]
    return {
        "org_full": ws["Q4"].value,
        "org_short": ws["A27"].value,
        "vehicle": ws["R9"].value,
        "plate": ws["AB11"].value,
        "driver": ws["M14"].value,
        "driver_return": ws["BQ41"].value,
        "driver_handover": ws["AD50"].value,
        "fuel_grade": ws["BG44"].value,
    }


def test_header_reflects_current_run_not_template_defaults(tmp_path):
    inp = _input_for("ООО «Ромашка»", "Петров П.П.", "Lada Vesta", "А000АА 77")
    res = generate_month_auto_seed(inp)
    out_path = write_waybills(
        template_path=BUILTIN_TEMPLATE, output_path=tmp_path / "wb.xlsx", out=res.output,
    )
    cells = _front_header_cells(out_path)

    assert cells["vehicle"] == "Lada Vesta"
    assert cells["plate"] == "А000АА 77"
    assert cells["driver"] == "Петров П.П."
    assert cells["driver_return"] == "Петров П.П."
    assert cells["driver_handover"] == "Петров П.П."
    assert cells["org_short"] == "ООО «Ромашка»"
    assert cells["fuel_grade"] == "АИ-92"
    assert "ОГРН 1234567890123" in cells["org_full"]

    # Ничего из старого builtin-шаблона (данные разработчика) не должно остаться.
    for value in cells.values():
        assert "Тихонов" not in value
        assert "Changan" not in value


def test_short_name_used_for_task_block_full_for_header(tmp_path):
    """Блок «Задание водителю» (A27) берёт сокращённое наименование, а шапка
    с реквизитами (Q4) — полное. Иначе длинное имя не вмещается в A27."""
    inp = _input_for("ООО «Ромашка»", "Петров П.П.", "Lada Vesta", "А000АА 77")
    inp.organization.name = (
        "Общество с ограниченной ответственностью «Ромашка и партнёры»"
    )
    inp.organization.short_name = "ООО «Ромашка»"
    res = generate_month_auto_seed(inp)
    out_path = write_waybills(
        template_path=BUILTIN_TEMPLATE, output_path=tmp_path / "wb.xlsx", out=res.output,
    )
    cells = _front_header_cells(out_path)
    assert cells["org_short"] == "ООО «Ромашка»"
    assert cells["org_full"].startswith("Общество с ограниченной ответственностью")


def test_header_differs_between_two_different_profiles(tmp_path):
    """Два разных прогона с разными профилями не должны давать одинаковую шапку —
    это и есть класс бага, который здесь фиксируется."""
    inp_a = _input_for("ООО «Альфа»", "Иванов И.И.", "Kia Rio", "А111АА 77")
    inp_b = _input_for("ООО «Бета»", "Сидоров С.С.", "Hyundai Solaris", "В222ВВ 77")

    out_a = write_waybills(
        template_path=BUILTIN_TEMPLATE, output_path=tmp_path / "a.xlsx",
        out=generate_month_auto_seed(inp_a).output,
    )
    out_b = write_waybills(
        template_path=BUILTIN_TEMPLATE, output_path=tmp_path / "b.xlsx",
        out=generate_month_auto_seed(inp_b).output,
    )

    cells_a = _front_header_cells(out_a)
    cells_b = _front_header_cells(out_b)
    assert cells_a != cells_b
