"""Integration-тесты FastAPI: новый многостраничный UX."""
from __future__ import annotations

import io
from pathlib import Path

import openpyxl
import pytest
from fastapi.testclient import TestClient

FIXTURES = Path(__file__).resolve().parent / "fixtures"

SETUP_DATA = {
    "organization_name": "ООО ИИС",
    "mechanic_name": "Павлов В.О.",
    "driver_full_name": "Тихонов А.Ю.",
    "driver_snils": "098-274-108 00",
    "driver_license_number": "99 19 №306940",
    "driver_license_issue_date": "2020-10-23",
    "vehicle_make_model": "Changan CS55plus",
    "vehicle_license_plate": "Н960ХА 198",
    "vehicle_fuel_grade": "АИ-95",
    "vehicle_tank_capacity_l": "50",
    "base_city": "Санкт-Петербург",
    "base_street_type": "улица",
    "base_street_name": "Репищева",
    "base_house_number": "10",
    "base_corpus": "",
    "vehicle_fuel_consumption_l_per_100km": "10.4",
    "start_odometer_km": "49898",
    "start_fuel_balance_l": "31.37",
    "start_date": "2026-03-31",
}

ROUTES_TO_ADD = [
    ("Санкт-Петербург, Комсомола 19", 12, False, 3.0),
    ("Санкт-Петербург, Марсово Поле 1", 12, False, 2.0),
    ("Санкт-Петербург, Смоленская 27", 27, True, 1.5),
    ("Санкт-Петербург, Московское шоссе, 304", 44, True, 1.5),
    ("Санкт-Петербург, Новое шоссе, 17", 35, True, 2.0),
]


@pytest.fixture()
def client(authed_client) -> TestClient:
    return authed_client


def _setup_and_routes(client: TestClient) -> None:
    r = client.post("/setup", data=SETUP_DATA, follow_redirects=False)
    assert r.status_code == 303
    for addr, km, is_large, weight in ROUTES_TO_ADD:
        data = {"address": addr, "km_one_way": str(km), "weight": str(weight)}
        if is_large:
            data["is_large"] = "true"
        r = client.post("/routes/add", data=data, follow_redirects=False)
        assert r.status_code == 303


def test_root_redirects_to_dashboard_when_logged_in(client: TestClient):
    # Залогиненный пользователь: / → /dashboard (а /dashboard сам уведёт на
    # /setup, пока профиль не заполнен).
    r = client.get("/", follow_redirects=False)
    assert r.status_code == 302
    assert r.headers["location"] == "/dashboard"
    r2 = client.get("/dashboard", follow_redirects=False)
    assert r2.status_code in (302, 303)
    assert r2.headers["location"] == "/setup"


def test_setup_get_renders_form(client: TestClient):
    r = client.get("/setup")
    assert r.status_code == 200
    assert "Настройка сервиса" in r.text
    assert 'name="organization_name"' in r.text


def test_setup_post_redirects_to_routes(client: TestClient):
    r = client.post("/setup", data=SETUP_DATA, follow_redirects=False)
    assert r.status_code == 303
    assert r.headers["location"] == "/routes"


def test_root_redirects_to_dashboard_after_setup(client: TestClient):
    # Залогиненный пользователь с заполненным профилем → /dashboard
    _setup_and_routes(client)
    r = client.get("/", follow_redirects=False)
    assert r.status_code == 302
    assert r.headers["location"] == "/dashboard"


def test_routes_crud(client: TestClient):
    _setup_and_routes(client)
    r = client.get("/routes")
    assert r.status_code == 200
    assert "Комсомола 19" in r.text
    assert "Смоленская 27" in r.text


def test_dashboard_shows_profile(client: TestClient):
    _setup_and_routes(client)
    r = client.get("/dashboard")
    assert r.status_code == 200
    assert "Тихонов А.Ю." in r.text
    assert "49898" in r.text  # одометр
    assert "Changan CS55plus" in r.text


def test_generate_full_pipeline(client: TestClient):
    _setup_and_routes(client)
    r = client.post(
        "/generate",
        data={
            "year": "2026", "month": "4",
            "fuelings_text": "15.04 47.63л по 69.37\n23.04 45.69л по 69.26",
        },
        follow_redirects=False,
    )
    assert r.status_code == 200, r.text[:500]
    assert "валидно" in r.text
    assert "Скачать" in r.text
    assert "/download/waybill/2026/4" in r.text


def test_generate_rejects_out_of_month_fueling(client: TestClient):
    # Заправка с датой из другого месяца раньше молча выпадала из расчёта.
    _setup_and_routes(client)
    r = client.post(
        "/generate",
        data={
            "year": "2026", "month": "4",
            "fueling_date": "2026-05-15",  # май, а генерируем апрель
            "fueling_liters": "47.63",
            "fueling_price": "69.37",
        },
        follow_redirects=False,
    )
    assert r.status_code == 400
    assert "15.05.2026" in r.text
    assert "Апрель" in r.text


def test_download_waybill_is_valid_xlsx(client: TestClient):
    _setup_and_routes(client)
    client.post("/generate", data={
        "year": "2026", "month": "4",
        "fuelings_text": "15.04 47.63л по 69.37\n23.04 45.69л по 69.26",
    })
    r = client.get("/download/waybill/2026/4")
    assert r.status_code == 200
    assert r.content[:4] == b"PK\x03\x04"
    wb = openpyxl.load_workbook(io.BytesIO(r.content), data_only=False)
    assert len(wb.sheetnames) == 22  # 11 пар × 2 листа
    ws = wb.worksheets[0]
    assert ws["AD2"].value == 1  # день 1 апреля
    assert ws["BU52"].value == 31.37  # стартовый остаток
    assert ws["BU31"].value == 49898  # стартовый одометр


def test_download_fuel_month_is_valid_xlsx(client: TestClient):
    _setup_and_routes(client)
    client.post("/generate", data={
        "year": "2026", "month": "4",
        "fuelings_text": "15.04 47.63л по 69.37\n23.04 45.69л по 69.26",
    })
    r = client.get("/download/fuel-month/2026/4")
    assert r.status_code == 200
    assert r.content[:4] == b"PK\x03\x04"
    wb = openpyxl.load_workbook(io.BytesIO(r.content), data_only=False)
    ws = wb.active
    # Должны быть заголовок Апрель 2026 и строка Итого
    has_header = False
    has_total = False
    for row in ws.iter_rows(values_only=True):
        for v in row:
            if v == "Апрель 2026":
                has_header = True
            if v == "Итого:":
                has_total = True
    assert has_header and has_total


def test_history_lists_runs(client: TestClient):
    _setup_and_routes(client)
    client.post("/generate", data={
        "year": "2026", "month": "4",
        "fuelings_text": "15.04 47.63л по 69.37\n23.04 45.69л по 69.26",
    })
    r = client.get("/history")
    assert r.status_code == 200
    assert "Апрель 2026" in r.text
    assert "путевые" in r.text
    assert "журнал" in r.text


def test_vehicle_state_updates_after_generation(client: TestClient):
    """После генерации текущий одометр и остаток на дашборде должны измениться."""
    _setup_and_routes(client)
    r = client.get("/dashboard")
    assert "49898" in r.text  # до генерации
    client.post("/generate", data={
        "year": "2026", "month": "4",
        "fuelings_text": "15.04 47.63л по 69.37\n23.04 45.69л по 69.26",
    })
    r = client.get("/dashboard")
    # Одометр должен вырасти после месяца поездок
    from putevoy.storage.repo import get_profile
    prof = get_profile()
    assert prof["state"]["current_odometer_km"] > 49898


def test_regenerate_same_month_idempotent(client: TestClient):
    """Повторная генерация того же месяца не должна дрейфовать состояние."""
    _setup_and_routes(client)
    client.post("/generate", data={
        "year": "2026", "month": "4",
        "fuelings_text": "15.04 47.63л по 69.37\n23.04 45.69л по 69.26",
    })
    from putevoy.storage.repo import get_profile, list_runs
    state_after_first = get_profile()["state"]
    assert len(list_runs()) == 1

    # Перегенерируем тот же месяц
    client.post("/generate", data={
        "year": "2026", "month": "4",
        "fuelings_text": "15.04 47.63л по 69.37\n23.04 45.69л по 69.26",
    })
    state_after_second = get_profile()["state"]
    assert len(list_runs()) == 1  # не плодим дублей
    # Состояние должно быть таким же (один и тот же seed=1 даёт ту же раскладку)
    assert state_after_first == state_after_second
