"""Golden-тесты для марта и февраля 2026 — реальные исторические месяцы."""
from __future__ import annotations

from pathlib import Path

import openpyxl
import pytest

from putevoy.generator.generate import generate_month_auto_seed
from putevoy.generator.models import MonthlyInput
from putevoy.generator.validators import validate
from putevoy.generator.writers.fuel_log_writer import append_month
from putevoy.generator.writers.waybill_writer import write_waybills

FIXTURES = Path(__file__).resolve().parent / "fixtures"
EXAMPLES = Path(__file__).resolve().parents[1] / "examples"


@pytest.fixture(scope="module")
def march_run(tmp_path_factory):
    out_dir = tmp_path_factory.mktemp("march_2026")
    inp = MonthlyInput.model_validate_json(
        (EXAMPLES / "config_march_2026.json").read_text(encoding="utf-8")
    )
    res = generate_month_auto_seed(inp)
    waybill = write_waybills(
        template_path=FIXTURES / "03__Путевой_лист_март_2026.xlsx",
        output_path=out_dir / "waybill.xlsx", out=res.output,
    )
    fuel = append_month(
        template_path=FIXTURES / "Топливо_февраль_2026.xlsx",
        output_path=out_dir / "fuel.xlsx", out=res.output,
    )
    return res, waybill, fuel


@pytest.fixture(scope="module")
def february_run(tmp_path_factory):
    out_dir = tmp_path_factory.mktemp("february_2026")
    inp = MonthlyInput.model_validate_json(
        (EXAMPLES / "config_february_2026.json").read_text(encoding="utf-8")
    )
    res = generate_month_auto_seed(inp)
    waybill = write_waybills(
        template_path=FIXTURES / "02__Путевой_лист_февраль_2026.xlsx",
        output_path=out_dir / "waybill.xlsx", out=res.output,
    )
    return res, waybill


def test_march_auto_seed_passes_validation(march_run):
    res, *_ = march_run
    report = validate(res.output)
    critical = [i for i in report.issues if i.severity == "error"]
    assert not critical, [i.message for i in critical]


def test_march_correct_working_days_count(march_run):
    res, *_ = march_run
    # Март 2026: 31 день, праздник 9 марта (вс 8.03 → перенос на пн 9.03),
    # выходные 1, 7, 8, 14, 15, 21, 22, 28, 29 = 9
    # 31 - 9 - 1 (9 марта) = 21
    assert len(res.output.days) == 21


def test_march_waybill_sheets(march_run):
    _, waybill, _ = march_run
    wb = openpyxl.load_workbook(waybill, data_only=False)
    # 21 рабочий день: 10 пар + 1 одиночный лист = 21 sheets? Нет, 10 пар = 20 листов + 2 = 22 (пара = 2 листа)
    # 21 день = 10 пар + 1 одиночный → 10×2 + 2 = 22 листа (одиночный тоже двусторонний)
    assert len(wb.sheetnames) == 22
    # Последняя пара — одиночный день
    last_front = wb.sheetnames[-2]
    assert "," not in last_front, f"Последний лист должен быть одиночным днём, а не парой: {last_front}"


def test_february_auto_seed_passes_validation(february_run):
    res, _ = february_run
    report = validate(res.output)
    critical = [i for i in report.issues if i.severity == "error"]
    assert not critical, [i.message for i in critical]


def test_february_correct_working_days_count(february_run):
    res, _ = february_run
    # Февраль 2026: 28 дней, 23.02 (пн) — праздник
    # Выходные: 1, 7, 8, 14, 15, 21, 22, 28 = 8 → 28 - 8 - 1 = 19
    assert len(res.output.days) == 19
