"""Golden-тест: полная генерация апреля 2026 на реальных fixtures.

Проверяем структурные инварианты (а не точное содержимое, поскольку
seed определяет случайный выбор маршрутов):
- Количество и формат имён листов
- Заполненность ключевых ячеек лицевой
- Непрерывность одометра между днями и через формулу FS31=AA43
- Запись маршрутов в правильные ячейки оборотной
- Создание файла Топливо.xlsx с заголовком месяца и строкой Итого
"""
from __future__ import annotations

import json
from datetime import time
from pathlib import Path

import openpyxl
import pytest
from openpyxl.utils import column_index_from_string

from putevoy.generator.generate import generate_month
from putevoy.generator.models import MonthlyInput
from putevoy.generator.writers.fuel_log_writer import append_month
from putevoy.generator.writers.waybill_writer import write_waybills

FIXTURES = Path(__file__).resolve().parent / "fixtures"
CONFIG = Path(__file__).resolve().parents[1] / "examples" / "config_april_2026.json"


@pytest.fixture(scope="module")
def generated(tmp_path_factory) -> tuple[Path, Path]:
    out_dir = tmp_path_factory.mktemp("april_2026")
    inp = MonthlyInput.model_validate_json(CONFIG.read_text(encoding="utf-8"))
    monthly_out = generate_month(inp)
    waybill = write_waybills(
        template_path=FIXTURES / "04__Путевой_лист_апрель_2026.xlsx",
        output_path=out_dir / "waybill.xlsx",
        out=monthly_out,
    )
    # Используем март как шаблон: реалистично — на момент генерации апреля
    # у пользователя есть Топливо.xlsx, заканчивающийся мартом.
    fuel = append_month(
        template_path=FIXTURES / "Топливо_март_2026.xlsx",
        output_path=out_dir / "fuel.xlsx",
        out=monthly_out,
    )
    return waybill, fuel


def test_waybill_sheet_count(generated):
    wb = openpyxl.load_workbook(generated[0], data_only=False)
    # 22 рабочих дня → 11 пар → 22 листа
    assert len(wb.sheetnames) == 22


def test_waybill_sheet_naming(generated):
    wb = openpyxl.load_workbook(generated[0], data_only=False)
    # каждая чётная (0,2,4...) — лицевая, нечётная — оборотная с суффиксом " 2"
    for i in range(0, len(wb.sheetnames), 2):
        front, back = wb.sheetnames[i], wb.sheetnames[i + 1]
        assert back == front + " 2"
        assert "апр" in front


def test_waybill_first_pair_day_left_filled(generated):
    wb = openpyxl.load_workbook(generated[0], data_only=False)
    ws = wb["01,02 апр"]
    assert ws["AD2"].value == 1
    assert ws["AI2"].value == "апреля"
    assert ws["AX2"].value == "26"
    assert ws["BU31"].value == 49874  # carry_over одометр
    assert ws["BU52"].value == 31.37  # carry_over остаток
    assert ws["BP33"].value == "01.04.2026 09:00"


def test_waybill_first_pair_day_right_uses_formulas(generated):
    wb = openpyxl.load_workbook(generated[0], data_only=False)
    ws = wb["01,02 апр"]
    # День 2: одометр и остаток — формулы из дня 1
    assert ws["FS31"].value == "=AA43"
    assert ws["FS52"].value == "=BU53"


def test_waybill_back_routes_filled(generated):
    wb = openpyxl.load_workbook(generated[0], data_only=False)
    ws = wb["01,02 апр 2"]
    # День 1, правая половина: маршрут в r7 (outbound) и r8 (inbound)
    AK = column_index_from_string("AK")
    AP = column_index_from_string("AP")
    BC = column_index_from_string("BC")
    AU = column_index_from_string("AU")
    AY = column_index_from_string("AY")
    assert ws.cell(7, AK).value == "Санкт-Петербург, Репищева 10"
    assert ws.cell(8, AP).value == "Санкт-Петербург, Репищева 10"
    assert ws.cell(7, BC).value is not None
    assert ws.cell(7, AU).value == time(9, 0)
    assert ws.cell(8, AY).value == time(12, 0)


def test_fuel_log_has_month_header_and_total(generated):
    wb = openpyxl.load_workbook(generated[1], data_only=False)
    ws = wb.active
    found_header = False
    found_total = False
    for r in range(1, ws.max_row + 1):
        v = ws.cell(r, 4).value
        if v == "Апрель 2026":
            found_header = True
        if v == "Итого:":
            # Проверим, что в L формула SUM
            l_val = ws.cell(r, 12).value
            if isinstance(l_val, str) and l_val.startswith("=SUM"):
                found_total = True
    assert found_header, "Не найден заголовок месяца"
    assert found_total, "Не найдена строка Итого с SUM"


def test_fuel_log_appends_correct_number_of_rows(generated):
    """Топливо.xlsx должен содержать ровно столько строк за апрель,
    сколько у нас trip-объектов (учитывая что 1 поездка = 2 строки туда+обратно)."""
    inp = MonthlyInput.model_validate_json(CONFIG.read_text(encoding="utf-8"))
    monthly_out = generate_month(inp)
    expected_data_rows = sum(len(d.trips) for d in monthly_out.days)

    wb = openpyxl.load_workbook(generated[1], data_only=False)
    ws = wb.active
    # Найдём заголовок Апрель 2026, потом считаем строки между ним и Итого
    header_row = None
    total_row = None
    for r in range(1, ws.max_row + 1):
        v = ws.cell(r, 4).value
        if v == "Апрель 2026":
            header_row = r
        elif v == "Итого:" and header_row is not None and total_row is None:
            total_row = r
    assert header_row and total_row
    # Между заголовком и итого: одна пустая строка + N data строк
    data_rows = total_row - header_row - 2  # минус заголовок, минус пустая
    assert data_rows == expected_data_rows
